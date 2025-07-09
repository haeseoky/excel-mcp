#!/usr/bin/env python3
"""
Excel MCP Server for Claude Desktop
Provides Excel file manipulation capabilities through MCP protocol
"""

import asyncio
import json
import sys
import traceback
from typing import Any, Dict, List, Optional, Union
import pandas as pd
import openpyxl
from pathlib import Path
import logging

# MCP 프로토콜 구현
class MCPServer:
    def __init__(self):
        self.tools = {}
        self.resources = {}
        self.setup_logging()
        self.register_tools()
        
    def setup_logging(self):
        """로깅 설정"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('excel_mcp.log'),
                logging.StreamHandler(sys.stderr)
            ]
        )
        self.logger = logging.getLogger(__name__)

    def register_tools(self):
        """사용 가능한 도구들 등록"""
        self.tools = {
            "read_excel": {
                "name": "read_excel",
                "description": "Excel 파일을 읽어서 데이터를 반환합니다.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Excel 파일 경로"
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "시트 이름 (선택사항, 기본값: 첫 번째 시트)",
                            "default": None
                        },
                        "rows": {
                            "type": "integer",
                            "description": "읽을 행 수 제한 (선택사항)",
                            "default": None
                        }
                    },
                    "required": ["file_path"]
                }
            },
            "write_excel": {
                "name": "write_excel",
                "description": "데이터를 Excel 파일로 저장합니다.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "저장할 Excel 파일 경로"
                        },
                        "data": {
                            "type": "array",
                            "description": "저장할 데이터 (딕셔너리 배열)"
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "시트 이름",
                            "default": "Sheet1"
                        }
                    },
                    "required": ["file_path", "data"]
                }
            },
            "get_excel_info": {
                "name": "get_excel_info",
                "description": "Excel 파일의 기본 정보를 가져옵니다.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Excel 파일 경로"
                        }
                    },
                    "required": ["file_path"]
                }
            },
            "analyze_excel": {
                "name": "analyze_excel",
                "description": "Excel 데이터를 분석하여 통계 정보를 제공합니다.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Excel 파일 경로"
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "분석할 시트 이름",
                            "default": None
                        }
                    },
                    "required": ["file_path"]
                }
            },
            "filter_excel_data": {
                "name": "filter_excel_data",
                "description": "Excel 데이터를 필터링합니다.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Excel 파일 경로"
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "시트 이름",
                            "default": None
                        },
                        "filters": {
                            "type": "object",
                            "description": "필터 조건 (컬럼명: 값)"
                        }
                    },
                    "required": ["file_path", "filters"]
                }
            }
        }

    async def handle_message(self, message: Dict[str, Any]) -> Dict[str, Any]:
        """MCP 메시지 처리"""
        try:
            method = message.get("method")
            params = message.get("params", {})
            msg_id = message.get("id")

            if method == "initialize":
                return await self.handle_initialize(msg_id, params)
            elif method == "tools/list":
                return await self.handle_list_tools(msg_id)
            elif method == "tools/call":
                return await self.handle_call_tool(msg_id, params)
            else:
                return self.error_response(msg_id, -32601, f"Method not found: {method}")

        except Exception as e:
            self.logger.error(f"Error handling message: {e}\n{traceback.format_exc()}")
            return self.error_response(message.get("id"), -32603, str(e))

    async def handle_initialize(self, msg_id: int, params: Dict[str, Any]) -> Dict[str, Any]:
        """초기화 핸들러"""
        return {
            "jsonrpc": "2.0",
            "id": msg_id,
            "result": {
                "protocolVersion": "2024-11-05",
                "capabilities": {
                    "tools": {},
                    "resources": {}
                },
                "serverInfo": {
                    "name": "excel-mcp-server",
                    "version": "1.0.0"
                }
            }
        }

    async def handle_list_tools(self, msg_id: int) -> Dict[str, Any]:
        """도구 목록 반환"""
        return {
            "jsonrpc": "2.0",
            "id": msg_id,
            "result": {
                "tools": list(self.tools.values())
            }
        }

    async def handle_call_tool(self, msg_id: int, params: Dict[str, Any]) -> Dict[str, Any]:
        """도구 호출 처리"""
        tool_name = params.get("name")
        arguments = params.get("arguments", {})

        if tool_name not in self.tools:
            return self.error_response(msg_id, -32602, f"Unknown tool: {tool_name}")

        try:
            if tool_name == "read_excel":
                result = await self.read_excel(**arguments)
            elif tool_name == "write_excel":
                result = await self.write_excel(**arguments)
            elif tool_name == "get_excel_info":
                result = await self.get_excel_info(**arguments)
            elif tool_name == "analyze_excel":
                result = await self.analyze_excel(**arguments)
            elif tool_name == "filter_excel_data":
                result = await self.filter_excel_data(**arguments)
            else:
                return self.error_response(msg_id, -32602, f"Tool not implemented: {tool_name}")

            return {
                "jsonrpc": "2.0",
                "id": msg_id,
                "result": {
                    "content": [
                        {
                            "type": "text",
                            "text": json.dumps(result, ensure_ascii=False, indent=2)
                        }
                    ]
                }
            }

        except Exception as e:
            self.logger.error(f"Error calling tool {tool_name}: {e}\n{traceback.format_exc()}")
            return self.error_response(msg_id, -32603, str(e))

    def error_response(self, msg_id: int, code: int, message: str) -> Dict[str, Any]:
        """에러 응답 생성"""
        return {
            "jsonrpc": "2.0",
            "id": msg_id,
            "error": {
                "code": code,
                "message": message
            }
        }

    # Excel 처리 메서드들
    async def read_excel(self, file_path: str, sheet_name: Optional[str] = None, rows: Optional[int] = None) -> Dict[str, Any]:
        """Excel 파일 읽기"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

            # pandas로 Excel 읽기
            kwargs = {}
            if sheet_name:
                kwargs['sheet_name'] = sheet_name
            if rows:
                kwargs['nrows'] = rows

            df = pd.read_excel(file_path, **kwargs)
            
            return {
                "success": True,
                "data": df.to_dict('records'),
                "shape": df.shape,
                "columns": df.columns.tolist(),
                "file_path": str(file_path)
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "file_path": str(file_path)
            }

    async def write_excel(self, file_path: str, data: List[Dict], sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """Excel 파일 쓰기"""
        try:
            df = pd.DataFrame(data)
            file_path = Path(file_path)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            df.to_excel(file_path, sheet_name=sheet_name, index=False)
            
            return {
                "success": True,
                "message": f"파일이 성공적으로 저장되었습니다: {file_path}",
                "rows_written": len(data),
                "file_path": str(file_path)
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "file_path": str(file_path)
            }

    async def get_excel_info(self, file_path: str) -> Dict[str, Any]:
        """Excel 파일 정보 가져오기"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

            # openpyxl로 시트 정보 가져오기
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets_info = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheets_info.append({
                    "name": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "dimensions": f"{sheet.max_column}x{sheet.max_row}"
                })
            
            workbook.close()
            
            return {
                "success": True,
                "file_path": str(file_path),
                "file_size": file_path.stat().st_size,
                "sheets": sheets_info,
                "total_sheets": len(sheets_info)
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "file_path": str(file_path)
            }

    async def analyze_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Excel 데이터 분석"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

            kwargs = {}
            if sheet_name:
                kwargs['sheet_name'] = sheet_name

            df = pd.read_excel(file_path, **kwargs)
            
            # 기본 통계 정보
            analysis = {
                "success": True,
                "file_path": str(file_path),
                "shape": df.shape,
                "columns": df.columns.tolist(),
                "data_types": {k: str(v) for k, v in df.dtypes.to_dict().items()},
                "missing_values": df.isnull().sum().to_dict(),
                "memory_usage": df.memory_usage(deep=True).to_dict()
            }
            
            # 숫자 컬럼 통계
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                analysis["numeric_statistics"] = df[numeric_cols].describe().to_dict()
            
            # 텍스트 컬럼 정보
            text_cols = df.select_dtypes(include=['object']).columns
            if len(text_cols) > 0:
                text_info = {}
                for col in text_cols:
                    text_info[col] = {
                        "unique_values": df[col].nunique(),
                        "most_common": df[col].value_counts().head(5).to_dict()
                    }
                analysis["text_statistics"] = text_info
            
            return analysis

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "file_path": str(file_path)
            }

    async def filter_excel_data(self, file_path: str, filters: Dict[str, Any], sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Excel 데이터 필터링"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

            kwargs = {}
            if sheet_name:
                kwargs['sheet_name'] = sheet_name

            df = pd.read_excel(file_path, **kwargs)
            
            # 필터 적용
            filtered_df = df.copy()
            for column, value in filters.items():
                if column in df.columns:
                    if isinstance(value, str):
                        # 문자열 포함 검색
                        filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(value, na=False)]
                    else:
                        # 정확한 값 매칭
                        filtered_df = filtered_df[filtered_df[column] == value]
            
            return {
                "success": True,
                "original_rows": len(df),
                "filtered_rows": len(filtered_df),
                "filters_applied": filters,
                "data": filtered_df.to_dict('records'),
                "file_path": str(file_path)
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "file_path": str(file_path)
            }

async def main():
    """메인 함수 - stdin/stdout으로 MCP 통신"""
    server = MCPServer()
    
    # stdin에서 JSON-RPC 메시지 읽기
    while True:
        try:
            line = sys.stdin.readline()
            if not line:
                break
                
            message = json.loads(line.strip())
            response = await server.handle_message(message)
            
            # stdout으로 응답 전송
            print(json.dumps(response), flush=True)
            
        except json.JSONDecodeError:
            continue
        except Exception as e:
            server.logger.error(f"Main loop error: {e}")
            break

if __name__ == "__main__":
    asyncio.run(main())
